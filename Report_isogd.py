from openpyxl import load_workbook
from PyQt5 import QtCore, QtWidgets

app = QtWidgets.QApplication([])

# file_path = QtWidgets.QFileDialog.getOpenFileName()[0]
# file_finish_path = QtWidgets.QFileDialog.getOpenFileName()[0]
file_path = 'C:/Users/18098/Downloads/2023.xlsx'
file_finish_path = 'D:/project_Python/Report_usluga/Отчет.xlsx'
wb = load_workbook(file_path)
wb_report = load_workbook(file_finish_path)
sheet = wb.active
sheet_report = wb_report.active
max_row = sheet.max_row

def usluga(tip):
	code_usluga = "9"
	if tip == "на присвоение адреса":
		code_usluga = '1'
	if tip == "на предоставление сведений ИСОГД":
		code_usluga = '3'
	if tip == "на выдачу копий архивных документов":
		code_usluga = '2'
	return code_usluga

def fiz_ur(subject):
	name = subject.split(' ')
	if ('"' not in subject) and (name[0] != '') and (name[1] != '') and (name[2] != '') \
			and len(subject.split()) == 3 \
			and name[0][0].isupper() and name[1][0].isupper() and name[2][0].isupper():
		a = '1'
	else:
		a = '2'
	return a

def kom_mfc_gosuslugi(prim):
	rez = "1"
	organ = "1"
	if type(prim) is str:
		prim = prim.lower()
		if ("мфц" in prim):
			organ = "2"
		if ("осуслу" in prim):
			organ = "3"
		if ("тказ" in prim):
			rez = "0"
	return organ, rez

# адрес - 1, архив - 2, ИСОГД - 3, не у слуга -9
	# физ - 1, юр - 2
	# комитет - 1, мфц - 2, госуслуги - 3
	# отказ - 0, положит - 1
list1 = []
list2 = []
for i in range(2, max_row + 1):
	cod_usluga = usluga(sheet[i][3].value)
	subject = fiz_ur(sheet[i][6].value)
	organ, rez = kom_mfc_gosuslugi(sheet[i][8].value)
	data1 = cod_usluga + subject + organ
	data2 = cod_usluga + subject + rez
	list1.append(data1)
	list2.append(data2)

	print(sheet[i][6].value)

sheet_report[5][3].value = list1.count('111')
sheet_report[7][3].value = list1.count('112')
sheet_report[9][3].value = list1.count('113')

sheet_report[5][4].value = list1.count('121')
sheet_report[7][4].value = list1.count('122')
sheet_report[9][4].value = list1.count('123')

sheet_report[5][5].value = list1.count('211')
sheet_report[7][5].value = list1.count('212')
sheet_report[9][5].value = list1.count('213')

sheet_report[5][6].value = list1.count('221')
sheet_report[7][6].value = list1.count('222')
sheet_report[9][6].value = list1.count('223')

sheet_report[5][7].value = list1.count('311')
sheet_report[7][7].value = list1.count('312')
sheet_report[9][7].value = list1.count('313')

sheet_report[5][8].value = list1.count('321')
sheet_report[7][8].value = list1.count('322')
sheet_report[9][8].value = list1.count('323')

sheet_report[16][3].value = list2.count('110')
sheet_report[16][4].value = list2.count('120')
sheet_report[16][5].value = list2.count('210')
sheet_report[16][6].value = list2.count('220')
sheet_report[16][7].value = list2.count('310')
sheet_report[16][8].value = list2.count('320')

print(list1)
print(list2)

wb_report.save(file_finish_path)
